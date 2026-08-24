import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

HEADERS = ["Nombre", "Cargo", "Correo", "Número de teléfono", "URL del perfil"]
COLUMN_WIDTHS = [35, 45, 30, 20, 55]


def sanitize_filename(value):
    value = re.sub(r'[\\/:*?"<>|]', "_", value).strip()
    return value or "empresa"


def export_to_excel(rows, company):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="0B66C2")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.get("name", ""),
            row.get("role", ""),
            "",
            "",
            row.get("url", ""),
        ])

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"empleados_{sanitize_filename(company)}_{stamp}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath, filename